import io
from flask import Blueprint, render_template, redirect, url_for, flash, send_file
from flask_login import login_required, current_user
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from app import db
from app.models import Factura, Auditoria

exportacion = Blueprint('exportacion', __name__)

def generar_excel(facturas):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Facturas'
    
    encabezados = [
        'Tipo comprobante',
        'Numero',
        'Fecha',
        'NIT Proveedor',
        'Proveedor',
        'Descripción',
        'Cantidad',
        'Valor unitario',
        'IVA',
        'Total',
        'Metodo pago',
        'Moneda'
    ]

    estilo_encabezado = Font(bold=True, color='FFFFFF')
    relleno_encabezado = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')

    for col, encabezado in enumerate(encabezados, 1):
        celda = ws.cell(row=1, column=col, value=encabezado)
        celda.font = estilo_encabezado
        celda.fill = relleno_encabezado
        celda.alignment = Alignment(horizontal='center')

    fila_actual = 2
    for factura in facturas:
        if factura.items:
            for item in factura.items:
                ws.cell(row=fila_actual, column=1, value=factura.tipo_factura or 'COMPRA')
                ws.cell(row=fila_actual, column=2, value=factura.numero_factura)
                ws.cell(row=fila_actual, column=3, value=str(factura.fecha_emision) if factura.fecha_emision else '')
                ws.cell(row=fila_actual, column=4, value=factura.nit_proveedor)
                ws.cell(row=fila_actual, column=5, value=factura.proveedor)
                ws.cell(row=fila_actual, column=6, value=item.descripcion)
                ws.cell(row=fila_actual, column=7, value=float(item.cantidad or 0))
                ws.cell(row=fila_actual, column=8, value=float(item.valor_unitario or 0))
                ws.cell(row=fila_actual, column=9, value=float(item.porcentaje_impuesto or 0))
                ws.cell(row=fila_actual, column=10, value=float(item.valor_total or 0))
                ws.cell(row=fila_actual, column=11, value=factura.metodo_pago or 'Credito')
                ws.cell(row=fila_actual, column=12, value=factura.moneda or 'COP')
                fila_actual += 1
        
        else:
            ws.cell(row=fila_actual, column=1, value=factura.tipo_factura or 'COMPRA')
            ws.cell(row=fila_actual, column=2, value=factura.numero_factura)
            ws.cell(row=fila_actual, column=3, value=str(factura.fecha_emision) if factura.fecha_emision else '')
            ws.cell(row=fila_actual, column=4, value=factura.nit_proveedor)
            ws.cell(row=fila_actual, column=5, value=factura.proveedor)
            ws.cell(row=fila_actual, column=6, value='')
            ws.cell(row=fila_actual, column=7, value=0)
            ws.cell(row=fila_actual, column=8, value=0)
            ws.cell(row=fila_actual, column=9, value=0)
            ws.cell(row=fila_actual, column=10, value=float(factura.valor_total or 0))
            ws.cell(row=fila_actual, column=11, value=factura.metodo_pago or 'Credito')
            ws.cell(row=fila_actual, column=12, value=factura.moneda or 'COP')
            fila_actual += 1

    for col in range(1, len(encabezados) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 20

    archivo_excel = io.BytesIO()
    wb.save(archivo_excel)
    archivo_excel.seek(0)
    return archivo_excel

@exportacion.route('/exportacion')
@login_required
def ver_exportacion():
    facturas_validadas = Factura.query.filter_by(estado = 'VALIDADA').order_by(Factura.fecha_validacion.desc()).all()
    return render_template('exportacion.html', facturas=facturas_validadas)

@exportacion.route('/exportacion/descargar', methods=['POST'])
@login_required
def descargar_exportacion():
    facturas = Factura.query.filter_by(
        estado='VALIDADA').all()

    if not facturas:
        flash('No hay facturas validadas para exportar.','warning')
        return redirect(
            url_for('exportacion.ver_exportacion'))

    try:
        archivo_excel = generar_excel(facturas)

        for factura in facturas:
            factura.estado = 'EXPORTADA'
            factura.fecha_exportacion = datetime.utcnow()
            db.session.add(Auditoria(
                id_usuario=current_user.id_usuario,
                accion='EXPORTACION_EXCEL',
                tabla_afectada='factura',
                id_referencia=factura.id_factura,
                detalles=(f'Factura {factura.numero_factura} ' f'exportada a Excel para carga en SIIGO')
            ))

        db.session.commit()

        flash('Facturas exportadas correctamente. ' 'Importe el archivo en SIIGO NUBE usando ' 'la opción de carga masiva.', 'success')

        nombre_archivo = (
            f'quantum_export_'
            f'{datetime.now().strftime("%Y%m%d_%H%M%S")}'
            f'.xlsx')

        return send_file(
            archivo_excel,
            mimetype=(
                'application/vnd.openxmlformats-'
                'officedocument.spreadsheetml.sheet'),
            as_attachment=True,
            download_name=nombre_archivo)

    except Exception as e:
        db.session.rollback()
        print(f'Error generando exportacion: {str(e)}')
        flash('Ocurrio un error al generar el archivo.', 'danger')
        return redirect(
            url_for('exportacion.ver_exportacion'))
